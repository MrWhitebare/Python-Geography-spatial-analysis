#EletricForm.py 使用电子表格
import shapefile
import openpyxl
#打开电子表格
xlsx=openpyxl.open(r"D:\Program Files\Python学习文档\samples\NYC_MUSEUMS_GEO\NYC_MUSEUMS_GEO.xlsx")
sheet=xlsx.active
sheet.title="MyForm"
writer=shapefile.Writer(r"D:\Program Files\Python学习文档\samples\NYC_MUSEUMS_GEO\Form.shp",shapefile.POINT)
for title in sheet.iter_rows(max_row=1,min_col=1,max_col=9):
    #固定范围为标题行
    for cell in title:
        writer.field(str(cell.value),'C',40)
for i in range(2,int(sheet.max_row)):
    #跳过第一行,索引从1开始
    values=list()
    for j in range(1,int(sheet.max_column)+1):
        #由于row,column 从一开始因此给列数+1
        values.append(sheet.cell(i,j).value)
    writer.record(*values)
    writer.point(float(values[-2]),float(values[-1]))
writer.close()
print("使用openpyxl完成！")
